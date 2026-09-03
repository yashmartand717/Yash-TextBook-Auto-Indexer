import os
os.environ['NO_PROXY'] = '*'

import streamlit as st

# 🚨 CRITICAL FIX: set_page_config MUST be the very first Streamlit command executed!
st.set_page_config(page_title="Document Data Extractor", layout="wide")

import pymupdf
import pandas as pd
import json
import io
import zipfile
import re
import time
import gc
import tempfile
import shutil
import openpyxl  # <-- Added this critical import!
from dotenv import load_dotenv
from openai import OpenAI
import httpx

# --- 1. CONFIGURATION & AUTHENTICATION ---
load_dotenv()

OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
try:
    if st.secrets.get("OPENAI_API_KEY"):
        OPENAI_API_KEY = st.secrets.get("OPENAI_API_KEY")
except Exception:
    pass

if OPENAI_API_KEY:
    try:
        custom_http = httpx.Client(proxy=None, timeout=60.0)
        openai_client = OpenAI(
            api_key=OPENAI_API_KEY, 
            http_client=custom_http,
            max_retries=3
        )
    except Exception as e:
        openai_client = None
        st.error(f"OpenAI Initialization Error: {e}")
else:
    openai_client = None


# --- 2. TEXT EXTRACTION & DISK-QUEUE PIPELINE ---
def natural_sort_key(s):
    return [int(text) if text.isdigit() else text.lower() for text in re.split(r'(\d+)', s)]

def build_disk_queue(uploaded_files, temp_dir):
    """Extracts and writes files directly to disk to keep RAM usage minimal."""
    queue = []
    for file in uploaded_files:
        filename = file.name.lower()
        if filename.endswith(".zip"):
            try:
                with zipfile.ZipFile(file) as zf:
                    valid_pdf_names = [
                        name for name in zf.namelist() 
                        if name.lower().endswith(".pdf") 
                        and not name.startswith("__MACOSX/") 
                        and not name.split("/")[-1].startswith("._")
                    ]
                    valid_pdf_names.sort(key=natural_sort_key)
                    for name in valid_pdf_names:
                        extracted_path = zf.extract(name, temp_dir)
                        queue.append(extracted_path)
            except Exception as e:
                st.error(f"Error extracting ZIP file {file.name}: {e}")
        elif filename.endswith(".pdf"):
            temp_path = os.path.join(temp_dir, file.name)
            with open(temp_path, "wb") as f:
                f.write(file.getbuffer())
            queue.append(temp_path)
            
    queue.sort(key=lambda p: natural_sort_key(os.path.basename(p)))
    return queue

def extract_text_from_pdf(pdf_path):
    """Reads a single PDF directly from disk."""
    full_text = []
    with pymupdf.open(pdf_path) as doc:
        for i, page in enumerate(doc):
            text = page.get_text("text")
            if text:
                full_text.append(f"--- PAGE {i+1} ---\n" + text)
    return "\n\n".join(full_text)

# --- 3. AI PROMPT FUNCTIONS ---
def process_text_with_ai(raw_text, filename, max_retries=3):
    """Extracts textbook chapter hierarchy."""
    if not openai_client: return []
    prompt = f"""
    You are an expert curriculum and textbook indexing system. I am providing you with the text of a document.
    Source File Name: {filename}
    Your task is to extract the structural hierarchy into a clean JSON array formatted for an educational database.
    CRITICAL INSTRUCTIONS:
    1. DYNAMIC HIERARCHY: 
       - If the text is an ENTIRE BOOK, set the broad section (like "Part I") as the "MODULE", and chapters as "CHAPTER".
       - If the text is a SINGLE CHAPTER, set the main Chapter Title as the "MODULE", and subtopics as "CHAPTER".
    2. NEVER leave "CHAPTER" blank. 
    3. STRICTLY IGNORE introductory outlines (like "CHAPTER FOCUS", "Learning Objectives"), page headers, and activity boxes.
    
    Output STRICTLY a valid JSON array of objects. Format exactly like this:
    [
      {{
        "MODULE": "PART I - What should I eat?",
        "CHAPTER": "Chapter 1 - Eat food."
      }}
    ]
    Textbook Content:
    {raw_text}
    """
    for attempt in range(1, max_retries + 1):
        try:
            response = openai_client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[{"role": "system", "content": "You are a precise data extraction assistant. Output clean JSON."}, {"role": "user", "content": prompt}],
                temperature=0.1, max_tokens=8192 
            )
            content = response.choices[0].message.content.strip()
            if content.startswith("```json"): content = content.replace("```json", "", 1)
            if content.endswith("```"): content = content[:-3]
            return json.loads(content.strip())
        except Exception:
            if attempt == max_retries: return []
            time.sleep(3 * attempt)

def process_experiment_with_ai(raw_text, filename, max_retries=3):
    """Extracts hardcoded lab manual sections."""
    if not openai_client: return []
    prompt = f"""
    You are an expert science educator. I am providing you with the text of a laboratory manual or experiment guide.
    Source File Name: {filename}
    
    Your task is to extract the core details of the experiment into a specific, hardcoded JSON structure.
    
    CRITICAL INSTRUCTIONS:
    1. Identify the name of the experiment (e.g., "Hooke's Law"). Let's call this [Exp Name].
    2. Extract or summarize the text into exactly these 5 sections:
       - "Introduction of [Exp Name]"
       - "Aim of [Exp Name]"
       - "Setup [Exp Name]"
       - "Observation"
       - "Conclusion"
    3. Do not invent information. Condense the relevant parts of the manual into comprehensive paragraphs.
    
    Output STRICTLY a valid JSON array of objects. Format exactly like this:
    [
      {{
        "SECTION": "Introduction of Hooke's Law",
        "CONTENT": "Detailed text extracted from the document..."
      }},
      {{
        "SECTION": "Aim of Hooke's Law",
        "CONTENT": "Detailed text extracted from the document..."
      }}
    ]
    Document Content:
    {raw_text}
    """
    for attempt in range(1, max_retries + 1):
        try:
            response = openai_client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[{"role": "system", "content": "You are a precise data extraction assistant. Output clean JSON."}, {"role": "user", "content": prompt}],
                temperature=0.2, max_tokens=8192 
            )
            content = response.choices[0].message.content.strip()
            if content.startswith("```json"): content = content.replace("```json", "", 1)
            if content.endswith("```"): content = content[:-3]
            return json.loads(content.strip())
        except Exception:
            if attempt == max_retries: return []
            time.sleep(3 * attempt)

def summarize_index_with_ai(detailed_data, max_retries=3):
    """Groups granular micro-chapters into broader thematic modules (Textbook Mode only)."""
    if not openai_client: return []
    prompt = f"""
    You are an expert curriculum designer. I am providing you with a highly detailed, granular book index (JSON).
    Your task is to group these granular items into a condensed, summarized index.
    
    CRITICAL INSTRUCTIONS:
    1. PRESERVE THE MODULES: Do NOT change or combine different "MODULE" names. 
    2. SUMMARIZE WITHIN MODULES: For each "MODULE", group every 4 to 6 related granular items together into a broader, thematic "CHAPTER".
    3. Output STRICTLY a valid JSON array of objects.
    
    Format exactly like this:
    [
      {{
        "MODULE": "PART I - What should I eat?",
        "CHAPTER": "Rules 1-6: Defining Real Food vs. Processed Products"
      }}
    ]
    Granular Index Data:
    {json.dumps(detailed_data, indent=2)}
    """
    for attempt in range(1, max_retries + 1):
        try:
            response = openai_client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[{"role": "system", "content": "You are a precise data summarization assistant. Output clean JSON."}, {"role": "user", "content": prompt}],
                temperature=0.3, max_tokens=4096
            )
            content = response.choices[0].message.content.strip()
            if content.startswith("```json"): content = content.replace("```json", "", 1)
            if content.endswith("```"): content = content[:-3]
            return json.loads(content.strip())
        except Exception:
            if attempt == max_retries: return []
            time.sleep(3 * attempt)

# --- 4. STREAMLIT UI & MAIN PIPELINE ---
st.title("📄 Multi-Format Document Extractor")
st.markdown("Automated curriculum and lab manual parser mapped directly to structured Excel sheets.")

# UI Mode Toggle
processing_mode = st.radio("Select Processing Mode:", ["📚 Textbook Index (Standard)", "🧪 Lab Manual (Experiment Details)"])

subject_input = st.text_input("Subject/Course Name", value="", placeholder="e.g. Physics, Science, Maths...")
uploaded_files = st.file_uploader("Upload PDFs or ZIP files", type=["pdf", "zip"], accept_multiple_files=True)

# Only show the summarization option if we are in Textbook mode
generate_summary = False
if "Textbook" in processing_mode:
    generate_summary = st.checkbox("Also generate a Summarized Index (Groups 4-6 micro-chapters into broader themes)", value=True)

if uploaded_files and st.button("Extract Data & Generate Excel", type="primary"):
    temp_dir = tempfile.mkdtemp()
    file_queue = build_disk_queue(uploaded_files, temp_dir)
    
    master_data = []
    progress_bar = st.progress(0, text="Starting queued text extraction...")
    
    # 1. FILE EXTRACTION QUEUE
    for idx, filepath in enumerate(file_queue):
        filename = os.path.basename(filepath)
        progress_bar.progress((idx + 1) / len(file_queue), text=f"Processing `{filename}` ({idx + 1}/{len(file_queue)})...")
        
        raw_text = extract_text_from_pdf(filepath)
        
        try: os.remove(filepath)
        except OSError: pass
        gc.collect()
        
        if len(raw_text.strip()) < 50:
            st.warning(f"⚠️ `{filename}` has no selectable text. Skipping.")
            continue
            
        # Branch AI processing based on selected mode
        if "Textbook" in processing_mode:
            extracted_data = process_text_with_ai(raw_text, filename)
            if extracted_data:
                for item in extracted_data:
                    master_data.append({
                        "SUBJECT": subject_input,
                        "MODULE": item.get("MODULE", ""),
                        "CHAPTER": item.get("CHAPTER", "")
                    })
        else:
            extracted_data = process_experiment_with_ai(raw_text, filename)
            if extracted_data:
                for item in extracted_data:
                    master_data.append({
                        "SUBJECT": subject_input,
                        "FILE": filename,
                        "SECTION": item.get("SECTION", ""),
                        "CONTENT": item.get("CONTENT", "")
                    })
            
        if not extracted_data:
            st.warning(f"⚠️ No structural data could be extracted from `{filename}`. Skipping.")
            
        time.sleep(1.0)
    
    shutil.rmtree(temp_dir, ignore_errors=True)
    
    # 2. OPTIONAL SUMMARIZATION PASS (Textbook Mode Only)
    summary_data = []
    if generate_summary and master_data and "Textbook" in processing_mode:
        progress_bar.progress(1.0, text="Generating summarized thematic grouping...")
        raw_summary = summarize_index_with_ai(master_data)
        if raw_summary:
            for item in raw_summary:
                summary_data.append({
                    "SUBJECT": subject_input,
                    "MODULE": item.get("MODULE", ""),
                    "CHAPTER": item.get("CHAPTER", "")
                })
            
    progress_bar.empty()
    
    # 3. EXCEL WORKBOOK GENERATION & UI DISPLAY
    if not master_data:
        st.error("❌ Critical Failure: Could not extract any data.")
    else:
        df_detailed = pd.DataFrame(master_data)
        excel_buffer = io.BytesIO()
        
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            sheet_title = 'Detailed Index' if "Textbook" in processing_mode else 'Experiment Details'
            df_detailed.to_excel(writer, index=False, sheet_name=sheet_title)
            worksheet_det = writer.sheets[sheet_title]
            
            # Auto-adjust column widths
            from openpyxl.utils import get_column_letter
            for i, col in enumerate(df_detailed.columns):
                # For Lab Manuals, wrap text for the large "CONTENT" column
                if col == "CONTENT":
                    worksheet_det.column_dimensions[get_column_letter(i + 1)].width = 80
                    for cell in worksheet_det[get_column_letter(i + 1)]: 
                        cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
                else:
                    max_len = max(df_detailed[col].astype(str).map(len).max(), len(str(col))) + 3
                    worksheet_det.column_dimensions[get_column_letter(i + 1)].width = max_len
                
            # Sheet 2: Summarized Index (Only for Textbooks)
            if generate_summary and summary_data:
                df_summary = pd.DataFrame(summary_data)
                df_summary.to_excel(writer, index=False, sheet_name='Summarized Index')
                worksheet_sum = writer.sheets['Summarized Index']
                for i, col in enumerate(df_summary.columns):
                    max_len = max(df_summary[col].astype(str).map(len).max(), len(str(col))) + 3
                    worksheet_sum.column_dimensions[get_column_letter(i + 1)].width = max_len

        st.success(f"🎉 Complete! Processed {len(df_detailed)} total entries.")
        
        if generate_summary and summary_data:
            tab1, tab2 = st.tabs(["Detailed Index", "Summarized Index"])
            with tab1: st.dataframe(df_detailed)
            with tab2: st.dataframe(df_summary)
        else:
            st.dataframe(df_detailed)
        
        st.download_button(
            label="📥 Download Excel File",
            data=excel_buffer.getvalue(),
            file_name="Extracted_Data.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )